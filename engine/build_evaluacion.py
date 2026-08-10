# Genera "Evaluacion Tiendas.xlsx": por tienda, su nivel de estrellas, el detalle
# de los 5 criterios (checar/tachar) y la PROXIMA estrella + como lograrla.
# NO incluye segmentacion (eso vive en Segmentacion.xlsx). Una hoja por formato.
# Usa el ULTIMO MES CERRADO (no el proyectado).
#
# Uso: python3 build_evaluacion.py data.json params.json "Evaluacion Tiendas.xlsx" [MES_LABEL]
import json, sys, math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA   = sys.argv[1] if len(sys.argv) > 1 else "data.json"
PARAMS = sys.argv[2] if len(sys.argv) > 2 else "params.json"
OUT    = sys.argv[3] if len(sys.argv) > 3 else "Evaluacion Tiendas.xlsx"

DB = json.load(open(DATA))
try:
    P = json.load(open(PARAMS))
except Exception:
    P = {}
CT = P.get("conv_target") or {"WM": 0.20, "BA": 0.07, "SC": 0.15}

months = DB["meta"]["months"]
proj   = DB["meta"].get("projected")
mi     = (proj["mi"] - 1) if proj else len(months) - 1      # ultimo mes CERRADO
MES = {"01":"Ene","02":"Feb","03":"Mar","04":"Abr","05":"May","06":"Jun",
       "07":"Jul","08":"Ago","09":"Sep","10":"Oct","11":"Nov","12":"Dic"}
y, m = months[mi].split("-"); LABEL = MES[m] + y[2:]
MESES_RESTANTES = 12 - int(m)    # meses que quedan en el año calendario tras el último cierre

S = DB["stars"]; START = S["start_idx"]; DIC25 = S["dic25_idx"]; crit = S["crit"]
stores = DB["stores"]
byidx = {s["i"]: s for s in stores}

# historia por tienda: hist[si][mi] = [monto, ge, meta, eleg]
hist = {}
for r in DB["recs"]:
    si, rmi, monto, ge, meta, eleg = r
    hist.setdefault(si, {})[rmi] = [monto, ge, meta, eleg]

def critbits(si, mm):
    a = crit.get(str(si))
    if not a: return None
    k = mm - START
    if k < 0 or k >= len(a) or a[k] < 0: return None
    return [(a[k] >> i) & 1 for i in range(5)]

PROXNAME = ["Vende", "Crece", "Cumple", "Constancia", "Eficiencia"]

def proxima(si, mm):
    f = byidx[si]["f"]; H = hist.get(si, {})
    def r(i): return H.get(i)
    def S3(a, b):
        s = 0
        for i in range(a, b + 1):
            rr = r(i)
            if rr: s += max(rr[1], 0)
        return s
    ca = me = el = 0; hits = 0
    for i in range(mm - 11, mm + 1):
        rr = r(i)
        if rr:
            ca += max(rr[1], 0); me += (rr[2] or 0); el += max(rr[3] or 0, 0)
            if (rr[2] or 0) > 0 and rr[1] >= rr[2]: hits += 1
    if ca == 0: return None
    conv = (ca / el) if el else 0; thr = 0.8 * CT[f]
    qNow = S3(mm - 2, mm); qPrev = S3(mm - 14, mm - 12)
    prevpres = any(r(i) for i in range(mm - 14, mm - 11))
    isnew = not prevpres
    mr = MESES_RESTANTES  # meses que quedan en el año calendario
    sfx_mr = f"(restan {mr} mes{'es' if mr != 1 else ''} del año)"
    cands = []
    if ca < 12:
        cands.append((ca/12, 0, f"Vender {math.ceil(12-ca)} GE más acumulados en los próximos 12 meses"))
    if mm >= DIC25 and not isnew and qPrev > 0 and qNow < qPrev:
        cands.append((qNow/qPrev, 1, f"Vender {math.ceil(qPrev-qNow)} GE más este trimestre vs. mismo trimestre del año anterior"))
    if not (me > 0 and ca >= me):
        cands.append(((ca/me) if me else 0, 2,
            f"Vender {math.ceil(max(0, me-ca))} GE más para cubrir la meta anual acumulada {sfx_mr}"))
    if hits < 10:
        cands.append((hits/10, 3,
            f"Cumplir la meta mensual en {10-hits} mes(es) más de los {mr} que quedan del año"))
    if not (el > 0 and conv >= thr):
        cands.append(((conv/thr) if thr > 0 else 0, 4,
            f"Subir conversión de {conv*100:.1f}% a {thr*100:.1f}% antes de cerrar {LABEL}"))
    if not cands: return {"max": True}
    cands.sort(key=lambda x: -x[0])
    return {"name": PROXNAME[cands[0][1]], "how": cands[0][2]}

# ---------- construir el libro ----------
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor="1F4E79")
ctr = Alignment(horizontal="center", vertical="center"); lft = Alignment(horizontal="left", vertical="center")
OKF = Font(color="1E7B34", bold=True); NOF = Font(color="C0392B", bold=True)
FNAME = {"WM": "Walmart", "BA": "Bodega Aurrerá", "SC": "Sam's Club"}
SHORT = ["Vende", "Crece", "Cumple", "Constancia", "Eficiencia"]
HDR = ["Determinante", "Tienda", "Nivel"] + SHORT + ["Próxima estrella", "Cómo lograrla"]

wb = openpyxl.Workbook(); wb.remove(wb.active)
for f in ["WM", "BA", "SC"]:
    ws = wb.create_sheet(FNAME[f][:31])
    for c, h in enumerate(HDR, 1):
        cc = ws.cell(1, c, h); cc.font = hf; cc.fill = hfill; cc.alignment = ctr
    n = 0
    for s in stores:
        if s["f"] != f: continue
        b = critbits(s["i"], mi)
        if b is None: continue
        nivel = sum(b)
        pr = proxima(s["i"], mi)
        if pr is None:
            pname, phow = "—", "—"
        elif pr.get("max"):
            pname, phow = "Nivel máximo (5★)", "Mantener los 5 logros"
        else:
            pname, phow = pr["name"], pr["how"]
        ws.append([s["clave"], s["n"], nivel,
                   *["✓" if x else "✗" for x in b], pname, phow])
        rr = ws.max_row
        ws.cell(rr, 2).alignment = lft
        ws.cell(rr, 3).alignment = ctr
        for i in range(5):
            cell = ws.cell(rr, 4 + i); cell.alignment = ctr
            cell.font = OKF if b[i] else NOF
        ws.cell(rr, 9).alignment = lft; ws.cell(rr, 10).alignment = lft
        n += 1
    widths = [13, 28, 7, 10, 15, 12, 10, 12, 15, 52]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    print(f"{f}: {n} tiendas · hoja {FNAME[f]}")

# Leyenda
lg = wb.create_sheet("Leyenda")
rows = [
    [f"Evaluación de tiendas · nivel de estrellas y acciones · al cierre de {LABEL}"], [],
    ["Criterio (columna)", "Qué significa"],
    ["Vende",      "Vendió ≥12 GE acumuladas en los últimos 12 meses"],
    ["Crece",      "No decreció vs. el mismo período del año previo"],
    ["Cumple",     "Cumplió la suma de sus metas de los últimos 12 meses"],
    ["Constancia", "Cumplió su meta mensual en al menos 10 de los últimos 12 meses"],
    ["Eficiencia", "≥80% de la conversión objetivo (WM 20% · BA 7% · SC 15%)"],
    [],
    ["✓", "Cumple el criterio"],
    ["✗", "No lo cumple"],
    ["Nivel", "Cantidad de criterios cumplidos (0 a 5)"],
    [],
    ["Próxima estrella", "El criterio prioritario más cercano de alcanzar"],
    ["Cómo lograrla", "La acción concreta para ganar esa estrella"],
    [],
    ["Nota", "Este archivo NO contiene segmentación (distritos/territoriales); eso vive en Segmentación.xlsx."],
]
for rr in rows: lg.append(rr)
lg.cell(1, 1).font = Font(bold=True, size=12)
for rn in (3,): lg.cell(rn, 1).font = Font(bold=True); lg.cell(rn, 2).font = Font(bold=True)
lg.column_dimensions["A"].width = 18; lg.column_dimensions["B"].width = 80

wb.save(OUT)
print(f"Listo · {OUT} · mes {LABEL}")

# ---------- subir a Google Drive ----------
EVAL_DRIVE_ID = "1BPuuY5J0sAxl7eatSbejcSK-NbpDb13A"
BOT_TOKEN_PATH = "/Users/reymon/Desktop/Proyectos/avance_ventas_bot/token.json"
BOT_CRED_PATH  = "/Users/reymon/Desktop/Proyectos/avance_ventas_bot/credentials.json"

def subir_evaluacion(xlsx_path, file_id, token_path, cred_path):
    try:
        import json as _json
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
        with open(token_path)  as f: tok = _json.load(f)
        with open(cred_path)   as f: ins = _json.load(f).get("installed", {})
        creds = Credentials(
            token=tok.get("token"),
            refresh_token=tok.get("refresh_token"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=ins.get("client_id"),
            client_secret=ins.get("client_secret"),
            scopes=tok.get("scopes", ["https://www.googleapis.com/auth/drive.file"])
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            # guardar token renovado
            tok_data = {
                "token": creds.token, "refresh_token": creds.refresh_token,
                "token_uri": creds.token_uri, "client_id": creds.client_id,
                "client_secret": creds.client_secret, "scopes": list(creds.scopes or [])
            }
            with open(token_path, "w") as f: _json.dump(tok_data, f, indent=2)
        svc   = build("drive", "v3", credentials=creds)
        media = MediaFileUpload(
            xlsx_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=True
        )
        result = svc.files().update(fileId=file_id, media_body=media).execute()
        print(f"Drive OK · id={result.get('id')}")
    except Exception as e:
        print(f"Drive upload omitido: {e}")

subir_evaluacion(OUT, EVAL_DRIVE_ID, BOT_TOKEN_PATH, BOT_CRED_PATH)
