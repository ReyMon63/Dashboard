#!/usr/bin/env python3
# Orquestador del Visor GE — Opción 2 (autosuficiente).
#
# Fuente de meses CERRADOS: data_closed.json (en el repo). NO depende de los históricos.
# El mes EN CURSO se proyecta desde el Avance de Ventas.
#
# AUTO-CONGELADO DE MES (rollover):
#   - Cada corrida guarda en pending.json la "foto cruda" (sin proyectar) del mes en curso
#     tomada del Avance (que es acumulado del mes).
#   - Cuando el Avance "se reinicia" (su monto cae drásticamente respecto a la foto guardada),
#     significa que el usuario ya pasó al mes siguiente -> se CONGELA la última foto (el mes que
#     cerró) dentro de data_closed.json y se avanza al siguiente mes.  Sin intervención manual.
#
# Uso: python3 run.py [--dry] [--force]
import os, sys, json, re, shutil, subprocess, calendar, hashlib, datetime
import openpyxl

ENGINE = os.path.dirname(os.path.abspath(__file__))
REPO   = os.path.dirname(ENGINE)
INPUTS = os.path.join(REPO, "inputs")
WORK   = os.path.join(REPO, ".work")
DRY    = "--dry"       in sys.argv
FORCE  = "--force"     in sys.argv
NOPROJ = "--noproject" in sys.argv   # publica solo hasta el último mes cerrado (sin proyectar el mes en curso)
AUTOFREEZE = "--autofreeze" in sys.argv  # activa el auto-congelado de mes (rollover). Off por defecto (en diseño).
ROLL_DROP = 0.55   # si el monto del Avance cae por debajo de este % del guardado -> mes nuevo

SHEET={'BA':'BA','WM':'WM','SC':'SMC'}

def next_month(ym):
    y,m=map(int,ym.split('-')); m+=1
    if m>12: y+=1; m=1
    return f"{y:04d}-{m:02d}"

def _to_int(v):
    try: return int(v)
    except: return None

def avance_raw(path, stores):
    """Foto CRUDA (factor 1) del Avance: agg y recs por tienda, sin proyectar."""
    look={}
    for s in stores:
        cl=_to_int(s['clave']); look[(s['f'], cl if cl is not None else s['clave'])]=s['i']
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    agg={f:[0.0,0.0,0.0,0.0,0] for f in ['WM','BA','SC']}; recs=[]
    for f,sh in SHEET.items():
        for r in wb[sh].iter_rows(min_row=2, values_only=True):
            cl=_to_int(r[0])
            if cl is None: continue
            si=look.get((f,cl))
            if si is None: continue
            venta=r[1] or 0; eleg=r[2] or 0; plan=r[3] or 0; ge=r[4] or 0
            recs.append([si, round(venta,2), round(ge,2), round(plan,2), round(eleg,2)])
            agg[f][0]+=venta; agg[f][1]+=ge; agg[f][2]+=plan; agg[f][3]+=eleg
            if ge>0: agg[f][4]+=1
    wb.close()
    for f in agg: agg[f]=[round(agg[f][i],2) if i<4 else agg[f][4] for i in range(5)]
    monto=sum(agg[f][0] for f in agg)
    return {'agg':agg,'recs':recs,'monto':round(monto,2)}

def fingerprint():
    wb=openpyxl.load_workbook(os.path.join(INPUTS,"Avance_Ventas.xlsx"), data_only=True, read_only=True)
    fp={}
    for f,sh in SHEET.items():
        ws=wb[sh]; s=0.0; n=0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if _to_int(r[0]) is None: continue
            s+=(r[1] or 0)+(r[4] or 0); n+=1
        fp[f]=hashlib.md5(f"{n}:{round(s,2)}".encode()).hexdigest()[:12]
    wb.close(); return fp

def jload(p, default): return json.load(open(p)) if os.path.exists(p) else default
def load_state():   return jload(os.path.join(REPO,"state.json"),   {"published_date":"","fp":{}})
def load_pending(): return jload(os.path.join(REPO,"pending.json"),  None)

def run(cmd, cwd):
    r=subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
    if r.returncode!=0:
        print("ERROR en"," ".join(cmd)); print(r.stdout); print(r.stderr); sys.exit(1)
    last=[l for l in r.stdout.strip().splitlines() if l.strip()]
    if last: print("  ·", last[-1])
    return r.stdout

def build_index(work):
    visor=open(os.path.join(work,"Visor_GE_Walmart.html"),encoding="utf-8").read()
    frag=open(os.path.join(work,"gate_fragment.html"),encoding="utf-8").read()
    scr=open(os.path.join(work,"gate_script.html"),encoding="utf-8").read()
    hashes=json.load(open(os.path.join(work,"params.json")))["users"]
    scr=scr.replace("__HASHES__", json.dumps(hashes, ensure_ascii=False))
    bi=visor.rfind("<body"); m=re.match(r"<body[^>]*>", visor[bi:]); pos=bi+m.end()
    out=visor[:pos]+"\n"+frag+"\n"+visor[pos:]
    idx=out.rfind("</body>"); out=out[:idx]+scr+"\n"+out[idx:]
    open(os.path.join(work,"index.html"),"w",encoding="utf-8").write(out)
    return len(hashes)

def freeze_into_closed(dc, month, snap):
    """Agrega 'month' (con la foto cruda snap) como mes cerrado en data_closed."""
    mi=len(dc['meta']['months'])
    dc['meta']['months'].append(month)
    for f in ['WM','BA','SC']:
        a=snap['agg'][f]; dc['agg'][f].append([round(a[0]),round(a[1]),round(a[2]),round(a[3]),a[4]])
    for si,venta,ge,plan,eleg in snap['recs']:
        dc['recs'].append([si, mi, round(venta), round(ge), round(plan), round(eleg)])
    return dc

def cierre_fp(snap):
    base="|".join(f"{f}:{snap['agg'][f][0]}:{snap['agg'][f][1]}:{snap['agg'][f][4]}" for f in ['WM','BA','SC'])
    return hashlib.md5(base.encode()).hexdigest()[:12]

def cierre_complete(snap):
    # los 3 formatos deben traer venta y al menos una GE
    return all(snap['agg'][f][0]>0 and snap['agg'][f][4]>0 for f in ['WM','BA','SC'])

def empty_cierre(path):
    try:
        wb=openpyxl.load_workbook(path)
        for sh in wb.sheetnames:
            ws=wb[sh]
            if ws.max_row>1: ws.delete_rows(2, ws.max_row-1)
        wb.save(path); return True
    except Exception as e:
        print("  · aviso: no pude vaciar el Cierre:", e); return False

def main():
    today=datetime.date.today()
    dcp=os.path.join(REPO,"data_closed.json")
    dc=json.load(open(dcp)); stores=dc['stores']
    last_closed=dc['meta']['months'][-1]
    cur_month=next_month(last_closed)
    rolled=None
    cier_fp=None
    # ---- CIERRE EXPLICITO: archivo 'Cierre_Ventas.xlsx' con los 3 formatos ----
    cierre_src=os.path.join(INPUTS,"Cierre_Ventas.xlsx")
    if os.path.exists(cierre_src):
        snap_c=avance_raw(cierre_src, stores)
        cier_fp=cierre_fp(snap_c)
        st_prev=load_state()
        this_cal=f"{today.year:04d}-{today.month:02d}"
        if not cierre_complete(snap_c):
            print("Cierre presente pero incompleto (falta algún formato con datos). No congelo aún.")
        elif cur_month in dc['meta']['months']:
            print(f"{cur_month} ya está cerrado. Ignoro el Cierre.")
        elif cur_month>=this_cal:
            print(f"El mes en curso {cur_month} aún no termina; un 'Cierre' ahora sería de un mes previo ya cerrado. Lo ignoro.")
        elif st_prev.get('cierre_fp')==cier_fp and not FORCE:
            print("Este Cierre ya se procesó. Nada nuevo que congelar.")
        else:
            dc=freeze_into_closed(dc, cur_month, snap_c)
            json.dump(dc, open(dcp,"w"), ensure_ascii=False, separators=(',',':'))
            rolled=cur_month; last_closed=cur_month; cur_month=next_month(cur_month)
            print(f"OK Cierre detectado: {rolled} congelado en data_closed. Ahora en curso: {cur_month}.")
    if AUTOFREEZE:
        snap=avance_raw(os.path.join(INPUTS,"Avance_Ventas.xlsx"), stores)  # foto cruda del Avance actual
        pend=load_pending()
        # ---- AUTO-CONGELADO: ¿el Avance se reinició (arrancó el mes siguiente)? ----
        if pend and pend.get('month')==cur_month and pend.get('monto',0)>0 \
           and 0 < snap['monto'] < pend['monto']*ROLL_DROP:
            dc=freeze_into_closed(dc, cur_month, pend)      # congela el mes que cerró (foto previa)
            json.dump(dc, open(dcp,"w"), ensure_ascii=False, separators=(',',':'))
            rolled=cur_month; last_closed=cur_month; cur_month=next_month(cur_month); pend=None
            print(f"↺ Cierre automático detectado: {rolled} congelado. Ahora en curso: {cur_month}.")
        json.dump({'month':cur_month, **snap}, open(os.path.join(REPO,"pending.json"),"w"),
                  ensure_ascii=False, separators=(',',':'))

    y,mn=map(int,cur_month.split('-')); dim=calendar.monthrange(y,mn)[1]
    this_cal=f"{today.year:04d}-{today.month:02d}"
    # Los avances reportan el acumulado HASTA AYER (el día de hoy aún no cierra),
    # así que los días con datos del mes = hoy - 1. En meses ya pasados, asof = dim.
    if cur_month==this_cal:
        asof=today.day-1
    else:
        asof=dim
    noproj=NOPROJ or asof<1     # día 1 del mes: aún no hay días con datos -> no proyectar
    if asof<1: asof=dim
    print(f"Último cierre: {last_closed} · mes en curso: {cur_month} · corte día {asof}/{dim}")

    st=load_state(); fp=fingerprint()
    fresh=all(fp[f]!=st["fp"].get(f) for f in fp)
    already=(st["published_date"]==today.isoformat())
    if not FORCE:
        if already and not rolled:
            print("Ya publiqué hoy y no hubo cierre nuevo. Nada que hacer."); return
        if not fresh and not rolled:
            print("Aún no están frescos los 3 formatos. No publico todavía."); return
    print(("Cierre nuevo → " if rolled else "Marcador OK → ")+"construyo y publico.")

    if os.path.exists(WORK): shutil.rmtree(WORK)
    os.makedirs(WORK)
    for fn in os.listdir(ENGINE):
        src=os.path.join(ENGINE,fn)
        if fn!="run.py" and os.path.isfile(src): shutil.copy(src, os.path.join(WORK,fn))
    shutil.copy(dcp, os.path.join(WORK,"data.json"))                                    # base cerrada
    shutil.copy(os.path.join(INPUTS,"Avance_Ventas.xlsx"), os.path.join(WORK,"Avance_Ventas.xlsx"))
    shutil.copy(os.path.join(INPUTS,"Parametros.xlsx"),    os.path.join(WORK,"Parametros.xlsx"))
    shutil.copy(os.path.join(INPUTS,"Segmentacion.xlsx"),  os.path.join(WORK,"Segmentacion_usuario.xlsx"))
    for pkg,fn in [("chart.js","chart.umd.min.js"),("xlsx","xlsx.full.min.js")]:
        d=os.path.join(WORK,"node_modules",pkg,"dist"); os.makedirs(d, exist_ok=True)
        shutil.copy(os.path.join(WORK,fn), os.path.join(d,fn))

    print("Pipeline:")
    run(["python3","read_params.py","Parametros.xlsx"], WORK)      # parámetros (fuente de verdad)
    run(["python3","stars.py","data.json","params.json"], WORK)    # estrellas (meses cerrados)
    run(["python3","recompute_all.py"], WORK)                      # presupuesto + segmentación
    if noproj:
        print("  · sin proyección; visor hasta el último mes cerrado")
    else:
        run(["python3","daily_update.py",cur_month,str(asof)], WORK)   # proyección del mes en curso
    run(["python3","build_visor.py"], WORK)
    nusers=build_index(WORK)
    print(f"  · index.html armado · {nusers} usuarios en el gate")

    shutil.copy(os.path.join(WORK,"index.html"), os.path.join(REPO,"index.html"))
    st={"published_date":today.isoformat(),"fp":fp,"last_closed":last_closed,"cur_month":cur_month,"asof":asof,
        "cierre_fp": (cier_fp if rolled else load_state().get("cierre_fp",""))}
    json.dump(st, open(os.path.join(REPO,"state.json"),"w"), ensure_ascii=False, indent=1)
    if DRY:
        print("[--dry] index.html actualizado. No hago git push."); return
    addf=["index.html","state.json","data_closed.json"]+(["pending.json"] if os.path.exists(os.path.join(REPO,"pending.json")) else [])
    run(["git","add",*addf], REPO)
    run(["git","-c","user.name=Visor GE Bot","-c","user.email=ReyMon63@users.noreply.github.com",
         "commit","-q","-m",f"Actualización · {cur_month} corte {asof}/{dim}"+(f" · cerró {rolled}" if rolled else "")], REPO)
    tok=os.environ.get("VISOR_GH_TOKEN","").strip()
    if not tok:
        for p in ("~/.gh_token", os.path.join(REPO,".gh_token")):
            p=os.path.expanduser(p)
            if os.path.exists(p): tok=open(p).read().strip(); break
    if not tok:
        print("No hay token de GitHub. index.html quedó armado pero no lo publiqué."); return
    run(["git","push",f"https://x-access-token:{tok}@github.com/ReyMon63/Dashboard.git","main"], REPO)
    print("Publicado en GitHub Pages.")
    if rolled:
        hist=os.environ.get("HIST_DIR","").strip()
        if hist and os.path.isdir(hist):
            run(["python3","append_hist_month.py",rolled,hist,os.path.join(WORK,"data.json")], WORK)
        else:
            print("  · (HIST_DIR no definido; no actualicé los Históricos con estrellas)")
        src=os.environ.get("CIERRE_SRC","").strip()
        if src and os.path.exists(src):
            if empty_cierre(src): print("  · Cierre de Ventas vaciado (listo para el próximo mes).")
        try: os.remove(cierre_src)
        except: pass

if __name__=="__main__":
    main()
