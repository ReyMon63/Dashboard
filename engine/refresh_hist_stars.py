# Reconstruye SOLO las columnas de estrellas (E1..E5 + "Cambio nivel") en un archivo
# "Histórico X (con estrellas).xlsx", tomando los valores recalculados de data.json.
# NO toca las columnas de datos ni su orden. Empareja por clave (columna 1 = DET).
# Uso: python3 refresh_hist_stars.py "<archivo con estrellas.xlsx>" data.json SC
import json, sys, re, shutil
import openpyxl

FILE = sys.argv[1]
DATA = sys.argv[2] if len(sys.argv) > 2 else "data.json"
FMT  = sys.argv[3] if len(sys.argv) > 3 else "SC"

DB = json.load(open(DATA))
months = DB["meta"]["months"]
S = DB["stars"]; START = S["start_idx"]; DIC25 = S["dic25_idx"]
crit = S["crit"]; levels = S["levels"]
clave2si = {int(s["clave"]): s["i"] for s in DB["stores"] if s["f"] == FMT}

MESN = {"Ene":1,"Feb":2,"Mar":3,"Abr":4,"May":5,"Jun":6,
        "Jul":7,"Ago":8,"Sep":9,"Oct":10,"Nov":11,"Dic":12}
def sheet_mi(name):
    m = re.match(r"([A-Za-z]{3})\s?(\d{2})$", str(name).strip())
    if not m: return None
    mm = MESN.get(m.group(1).capitalize())
    if not mm: return None
    ym = f"20{m.group(2)}-{mm:02d}"
    return months.index(ym) if ym in months else None

def bits(si, k):
    a = crit.get(str(si))
    if not a or not (0 <= k < len(a)) or a[k] < 0: return None
    return [(a[k] >> i) & 1 for i in range(5)]
def lvl(si, k):
    a = levels.get(str(si))
    if not a or not (0 <= k < len(a)): return None
    return a[k] if a[k] >= 0 else None

shutil.copy(FILE, FILE + ".bak_stars")
wb = openpyxl.load_workbook(FILE)
upd = 0
for sh in wb.sheetnames:
    mi = sheet_mi(sh)
    if mi is None or mi < START: continue
    ws = wb[sh]; k = mi - START; prevk = k - 1
    for r in range(2, ws.max_row + 1):
        cl = ws.cell(r, 1).value
        try: cl = int(cl)
        except: continue
        si = clave2si.get(cl)
        b = bits(si, k) if si is not None else None
        if b is None:
            for i in range(5): ws.cell(r, 9 + i, None)
            ws.cell(r, 14, None); continue
        pb = bits(si, prevk) if (si is not None and prevk >= 0) else None
        for i in range(5):
            pend = (i == 1 and mi < DIC25)
            if pend:
                val = ""
            elif pb is None or (i == 1 and mi == DIC25):
                val = b[i]
            else:
                now, prev = b[i], pb[i]
                if now and prev: val = 1
                elif not now and not prev: val = 0
                elif now and not prev: val = "+1"
                else: val = -1
            ws.cell(r, 9 + i, val)
        L = lvl(si, k); Lp = lvl(si, prevk) if prevk >= 0 else None
        camb = "—" if (L is None or Lp is None) else ("Más" if L > Lp else ("Menos" if L < Lp else "Igual"))
        ws.cell(r, 14, camb)
    upd += 1
    print(f"  · {sh}: estrellas recalculadas")
wb.save(FILE)
print(f"Listo · {upd} hojas de mes actualizadas · {FILE}")
