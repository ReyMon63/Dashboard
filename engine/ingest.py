# Ingesta de históricos CERRADOS -> data.json (stores, recs, agg, months).
# Detecta automáticamente TODAS las hojas de mes presentes (auto-cierre de mes).
# No calcula estrellas/presupuesto/segmentación (eso viene después en el pipeline).
import openpyxl, re, unicodedata, json, sys, os
from collections import defaultdict

INDIR = sys.argv[1] if len(sys.argv)>1 else "."   # carpeta con Historico_WM/BA/SC.xlsx
OUT   = sys.argv[2] if len(sys.argv)>2 else "data.json"

def norm(s):
    if s is None: return ""
    s=str(s).strip().lower()
    s=''.join(c for c in unicodedata.normalize('NFD',s) if unicodedata.category(c)!='Mn')
    return re.sub(r'\s+',' ',s)

def field(h):
    n=norm(h)
    if n in ('det','det.'): return 'clave'
    if n in ('tda','club'): return 'nombre'
    if n=='ventas': return 'AMB'
    if n=='monto': return 'monto'
    if n in ('ppto','ptto.'): return 'meta'
    if n=='garantias extendidas': return 'cantidad'
    if n in ('eleg','elegibles'): return 'elegibles'
    return None

MES={'Ene':'01','Feb':'02','Mar':'03','Abr':'04','May':'05','Jun':'06',
     'Jul':'07','Ago':'08','Sep':'09','Oct':'10','Nov':'11','Dic':'12'}
def sheet_ym(name):
    m=re.match(r'([A-Za-z]{3})\s?(\d{2})$', str(name).strip())
    if not m or m.group(1).capitalize() not in MES: return None
    return f"20{m.group(2)}-{MES[m.group(1).capitalize()]}"

def num(x): return float(x) if isinstance(x,(int,float)) else 0.0

# 1) descubrir el universo de meses (unión de hojas válidas en los 3 formatos), ordenado
month_ym=set()
sheets_by_f={}
for k in ["WM","BA","SC"]:
    wb=openpyxl.load_workbook(os.path.join(INDIR,f"Historico_{k}.xlsx"), read_only=True, data_only=True)
    valid={s: sheet_ym(s) for s in wb.sheetnames if sheet_ym(s)}
    sheets_by_f[k]=valid
    month_ym|=set(valid.values())
    wb.close()
months=sorted(month_ym)                       # p.ej. 2024-01 .. 2026-06
midx={m:i for i,m in enumerate(months)}

stores={}; records=[]; warn=[]
for k in ["WM","BA","SC"]:
    wb=openpyxl.load_workbook(os.path.join(INDIR,f"Historico_{k}.xlsx"), read_only=True, data_only=True)
    for sh,ym in sheets_by_f[k].items():
        rows=list(wb[sh].iter_rows(values_only=True))
        hidx=None
        for i,r in enumerate(rows[:6]):
            if any(norm(x) in ('det','det.') for x in r): hidx=i; break
        if hidx is None: warn.append(f"{k}/{sh}: sin encabezado"); continue
        cmap={}
        for ci,h in enumerate(rows[hidx]):
            f=field(h)
            if f and f not in cmap: cmap[f]=ci
        if 'clave' not in cmap: warn.append(f"{k}/{sh}: sin columna clave"); continue
        for r in rows[hidx+1:]:
            clave=r[cmap['clave']]
            if clave is None or not isinstance(clave,(int,float)): continue
            clave=int(clave)
            nombre=r[cmap['nombre']] if 'nombre' in cmap and cmap['nombre']<len(r) else None
            nombre=(str(nombre).strip() if nombre is not None else "")
            if k=='SC': monto=num(r[cmap['monto']]); cant=num(r[cmap['AMB']])
            else:       monto=num(r[cmap['AMB']]);   cant=num(r[cmap['cantidad']])
            meta=num(r[cmap['meta']]) if 'meta' in cmap else 0.0
            eleg=num(r[cmap['elegibles']]) if 'elegibles' in cmap else 0.0
            key=(k,clave)
            if key not in stores or (nombre and not stores[key]['nombre']):
                stores[key]={'clave':clave,'nombre':nombre,'formato':k}
            records.append((k,ym,clave,round(monto,2),round(cant,2),round(meta,2),round(eleg,2)))
    wb.close()

slist=sorted(stores.values(), key=lambda x:(x['formato'],x['clave']))
sidx={(s['formato'],s['clave']):i for i,s in enumerate(slist)}
recs=[[sidx[(f,cl)], midx[ym], m, c, mt, e] for (f,ym,cl,m,c,mt,e) in records]

agg={f:{} for f in ['WM','BA','SC']}
for si,mi,m,c,mt,e in recs:
    f=slist[si]['formato']; d=agg[f].setdefault(mi,[0,0,0,0,0])
    d[0]+=m; d[1]+=c; d[2]+=mt; d[3]+=e
    if c>0: d[4]+=1
def series(f):
    z=[0,0,0,0,0]
    return [[round(agg[f].get(i,z)[0],2),round(agg[f].get(i,z)[1],2),round(agg[f].get(i,z)[2],2),
             round(agg[f].get(i,z)[3],2),agg[f].get(i,z)[4]] for i in range(len(months))]

out={
 'meta':{'formatos':{'WM':'Walmart','BA':'Bodega Aurrerá','SC':"Sam's Club"},'months':months},
 'stores':[{'i':i,'f':s['formato'],'clave':s['clave'],'n':s['nombre']} for i,s in enumerate(slist)],
 'agg':{f:series(f) for f in ['WM','BA','SC']},
 'recs':recs,
}
json.dump(out, open(OUT,"w"), ensure_ascii=False, separators=(',',':'))
print(f"Ingesta OK · {len(months)} meses ({months[0]}..{months[-1]}) · {len(slist)} tiendas · {len(recs)} recs · warns={len(warn)}")
if warn: print("  ", warn[:5])
