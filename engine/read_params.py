# Lee Parámetros Base.xlsx y extrae TODOS los parámetros a params.json (fuente de verdad)
import openpyxl, json, sys, re
FN = sys.argv[1] if len(sys.argv)>1 else "Parámetros Base.xlsx"
wb = openpyxl.load_workbook(FN, data_only=True)
def num(v, default=None):
    if v is None: return default
    if isinstance(v,(int,float)): return float(v)
    s=str(v).replace('%','').replace(',','').strip()
    try: return float(s)
    except: return default
def sheet_rows(name):
    if name not in wb.sheetnames: return []
    return [ [c for c in r] for r in wb[name].iter_rows(values_only=True) ]

P={"_source":FN}

# --- Conversión objetivo (WM/BA/SC) ---
conv={}
for r in sheet_rows("Conversión objetivo"):
    if not r or r[0] is None: continue
    lab=str(r[0])
    for f,key in [("WM",["walmart","wm"]),("BA",["bodega","ba"]),("SC",["sam","sc"])]:
        if any(k in lab.lower() for k in key):
            v=None
            for cell in r[1:]:
                v=num(cell)
                if v is not None: break
            if v is not None: conv[f]= v/100 if v>1 else v
P["conv_target"]=conv or {"WM":0.20,"BA":0.07,"SC":0.15}

# --- Meta mínima por formato (pisos) ---
floor={}
for r in sheet_rows("Meta mínima por formato"):
    if not r or r[0] is None: continue
    lab=str(r[0]).lower()
    val=num(r[1])
    for f,key in [("WM",["walmart","wm"]),("BA",["bodega","ba"]),("SC",["sam","sc"])]:
        if any(k in lab for k in key) and val: floor[f]=int(val)
P["floor"]=floor or {"WM":146,"BA":73,"SC":122}

# --- Implant por formato ---
implant={}
for r in sheet_rows("Implant por formato"):
    if not r or r[0] is None: continue
    lab=str(r[0]).lower()
    name=r[2] if len(r)>2 and r[2] else None
    for f,key in [("WM",["walmart","wm"]),("BA",["bodega","ba"]),("SC",["sam","sc"])]:
        if any(k in lab for k in key) and name: implant[f]=str(name).strip()
P["implant"]=implant or {"WM":"Luzmaría Ramírez","BA":"Clara Olivares","SC":"Daniela Pérez"}

# --- Usuarios (para el gate) ---
import hashlib
users={}
for r in sheet_rows("Usuarios")[1:]:
    if not r or r[0] is None: continue
    name=str(r[0]).strip(); pw=r[1]
    if name and pw and not name.lower().startswith("nota"):
        users[hashlib.sha256(str(pw).strip().encode()).hexdigest()]=name.split("(")[0].strip()
P["users"]=users

# --- Generales: growth, alpha, meses de evento ---
growth=None; alpha=None; events=[]
txt=" ".join(str(c) for r in sheet_rows("Generales") for c in r if c is not None)
mg=re.search(r'(\d+(?:\.\d+)?)\s*%.*crecim', txt, re.I) or re.search(r'crecim\w*.*?(\d+(?:\.\d+)?)\s*%', txt, re.I)
for r in sheet_rows("Generales"):
    if not r: continue
    lab=" ".join(str(c) for c in r if c is not None).lower()
    if 'crecim' in lab:
        for c in r:
            v=num(c)
            if v is not None: growth=v/100 if v>1 else v; break
    if 'alpha' in lab or 'suaviz' in lab or 'α' in lab:
        for c in r:
            v=num(c)
            if v is not None and v<=1: alpha=v; break
P["growth"]=growth if growth is not None else 0.10
P["alpha"]=alpha if alpha is not None else 0.45
P["event_months"]=[5,11]

json.dump(P, open("params.json","w"), ensure_ascii=False, indent=1)
print("== Parámetros leídos de", FN, "==")
for k in ["growth","alpha","event_months","conv_target","floor","implant"]:
    print(f"  {k}: {P[k]}")
print(f"  usuarios: {len(P['users'])} claves")
