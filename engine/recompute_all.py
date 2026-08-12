import json, statistics as st, unicodedata, re
from collections import defaultdict
import openpyxl
P=json.load(open("params.json"))
DB=json.load(open("data.json"))
stores=DB['stores']; recs=DB['recs']; months=DB['meta']['months']; NM=len(months)
byS=defaultdict(dict)
for si,mi,mo,ca,me,el in recs: byS[si][mi]=(mo,ca,me,el)
def fmt(si): return stores[si]['f']
def calm(mi): return int(months[mi].split('-')[1])
STARS=DB['stars']; START=STARS['start_idx']
def starAt(si,mi):
    a=STARS['levels'].get(str(si)); return a[mi-START] if a and 0<=mi-START<len(a) else -1
DIC25=months.index('2025-12')
Y2025={m:months.index(f'2025-{m:02d}') for m in range(1,13)}
G=P['growth']; ALPHA=P['alpha']; EVENT=set(P['event_months']); FLOOR=P['floor']; CONV=P['conv_target']
NORMAL=[m for m in range(1,13) if m not in EVENT]
TR=range(NM-12,NM)   # últimos 12 meses (ventana móvil = base del presupuesto)
TRcal={calm(mi):mi for mi in TR}   # mes-calendario (1..12) -> índice dentro de la ventana de 12m

# estacionalidad de formato (meses normales)
fs=defaultdict(lambda:defaultdict(list))
for si in byS:
    for mi,(mo,ca,me,el) in byS[si].items():
        if calm(mi) in NORMAL: fs[fmt(si)][calm(mi)].append(max(ca,0))
fmt_w={f:{m:(st.mean(fs[f][m]) if fs[f][m] else 1) for m in NORMAL} for f in ['WM','BA','SC']}
for f in fmt_w:
    s=sum(fmt_w[f].values()); fmt_w[f]={m:fmt_w[f][m]/s for m in NORMAL}
def store_w(si):
    hm={m:[] for m in NORMAL}
    for mi,(mo,ca,me,el) in byS[si].items():
        if calm(mi) in NORMAL: hm[calm(mi)].append(max(ca,0))
    avg={m:(st.mean(hm[m]) if hm[m] else None) for m in NORMAL}
    if all(v in (None,0) for v in avg.values()): base=dict(fmt_w[fmt(si)])
    else:
        pres=[v for v in avg.values() if v]; fill=st.mean(pres) if pres else 1
        avg={m:(avg[m] or fill) for m in NORMAL}; s=sum(avg.values()) or 1
        base={m:avg[m]/s for m in NORMAL}
    flat=1/len(NORMAL); w={m:ALPHA*base[m]+(1-ALPHA)*flat for m in NORMAL}
    s=sum(w.values()); return {m:w[m]/s for m in NORMAL}
def v3(si):
    f=fmt(si); tier=starAt(si,DIC25); tier=1 if tier<0 else tier
    a={m:(byS[si][TRcal[m]][1] if TRcal[m] in byS[si] else 0) for m in range(1,13)}   # base: ventana 12m móvil
    out={m:a[m]*(1+G) for m in EVENT}
    base_norm=sum(a[m] for m in NORMAL)*(1+G); w=store_w(si)
    for m in NORMAL: out[m]=base_norm*w[m]
    mmin=FLOOR[f]//12   # mínimo MENSUAL = piso anual / 12, entero truncado
    out={m:max(out[m],mmin) for m in range(1,13)}   # ningún mes por debajo del mínimo mensual
    return out,tier

budget3={'params':{'growth':G,'alpha':ALPHA,'event':sorted(EVENT),'floor':FLOOR,'conv_target':CONV},'stores':{'WM':[],'BA':[],'SC':[]}}
for si in byS:
    f=fmt(si)
    v2025=sum(byS[si][Y2025[m]][1] for m in range(1,13) if Y2025[m] in byS[si])
    base12=sum(byS[si][TRcal[m]][1] for m in range(1,13) if TRcal[m] in byS[si])
    if base12<=0: continue   # incluye tiendas nuevas (con venta en los últimos 12m aunque no en 2025)
    el=sum(byS[si][x][3] for x in TR if x in byS[si])
    pz=sum(byS[si][x][1] for x in TR if x in byS[si])
    cur=sum((byS[si][x][2] or 0) for x in TR if x in byS[si])
    mm,tier=v3(si); m12=[round(mm[m]) for m in range(1,13)]; ann=sum(m12)
    budget3['stores'][f].append([si,tier,round(v2025),round(el),round(pz),round(cur),ann,m12])
DB['budget3']=budget3

# --- segmentación con implant de params ---
wb=openpyxl.load_workbook("Segmentacion_usuario.xlsx", data_only=True)
IMPLANT=P['implant']
def to_int(v):
    try: return int(v)
    except: return None
dist={}; terr={}
for f in ['WM','BA','SC']:
    for r in wb['Distritos '+f].iter_rows(min_row=2,values_only=True):
        c=to_int(r[1])
        if c is not None and r[0] not in (None,''): dist[(f,c)]=str(r[0]).strip()
    for r in wb['Territoriales '+f].iter_rows(min_row=2,values_only=True):
        c=to_int(r[2])
        if c is not None and r[1] not in (None,''): terr[(f,c)]=str(r[1]).strip()
seg={}
for s in stores:
    f=s['f']; c=to_int(s['clave'])
    d=dist.get((f,c)) or 'Sin distrito'
    t=terr.get((f,c)) or IMPLANT[f]
    seg[str(s['i'])]={'d':d,'t':t}
DB['seg']=seg

json.dump(DB, open("data.json","w"), ensure_ascii=False, separators=(',',':'))
# resumen
for f in ['WM','BA','SC']:
    arr=budget3['stores'][f]; sug=sum(r[6] for r in arr); cur=sum(r[5] for r in arr)
    enpiso=sum(1 for r in arr if r[6]<=FLOOR[f]+0.5)
    print(f"{f}: piso {FLOOR[f]} · {len(arr)} tiendas · V3 {sug:,.0f} (actual {cur:,.0f}, {(sug/cur-1)*100:+.0f}%) · {enpiso} en piso")
tt=sorted(set(v['t'] for v in seg.values()))
print("Territoriales:", len(tt), "| Implants:", {f:IMPLANT[f] for f in IMPLANT})
