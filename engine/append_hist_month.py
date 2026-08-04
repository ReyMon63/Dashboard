# Agrega la hoja de UN mes recién cerrado a los 3 Históricos "(con estrellas)".
# Datos (formato por formato) + columnas E1..E5 + "Cambio nivel", tomando:
#   - datos y estrellas desde data.json (ya con stars, generado por el pipeline)
# Uso: python3 append_hist_month.py 2026-08 "<HIST_DIR>" data.json
import json, sys, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

MONTH   = sys.argv[1]
HISTDIR = sys.argv[2]
DATA    = sys.argv[3] if len(sys.argv) > 3 else "data.json"

DB = json.load(open(DATA))
months = DB['meta']['months']
if MONTH not in months:
    print(f"ERROR: {MONTH} no está en data.json"); sys.exit(2)
mi = months.index(MONTH)
stores = DB['stores']

S = DB['stars']; START = S['start_idx']; DIC25 = S['dic25_idx']
levels = S['levels']; crit = S['crit']

MES = {'01':'Ene','02':'Feb','03':'Mar','04':'Abr','05':'May','06':'Jun',
       '07':'Jul','08':'Ago','09':'Sep','10':'Oct','11':'Nov','12':'Dic'}
y, m = MONTH.split('-'); LABEL = MES[m] + y[2:]

# ---- datos del mes: si -> (monto$, ge, meta, eleg) ----
row = {}
for si, rmi, mo, ca, me, el in DB['recs']:
    if rmi == mi: row[si] = (mo, ca, me, el)

# ---- estrellas ----
def bits(si, k):
    a = crit.get(str(si))
    if not a or not (0 <= k < len(a)) or a[k] < 0: return None
    return [(a[k] >> i) & 1 for i in range(5)]
def lvl(si, k):
    a = levels.get(str(si))
    if not a or not (0 <= k < len(a)): return None
    return a[k] if a[k] >= 0 else None

k = mi - START; prevk = k - 1

hf = Font(bold=True, color='FFFFFF'); hfill = PatternFill('solid', fgColor='2A2A28')
ctr = Alignment(horizontal='center')

HDR = {'WM': ['DET','TDA','VENTAS  ',' PPTO','Garantías Extendidas','Alcance Plan  ',' ELEG','Conv. '],
       'BA': ['DET','TDA','VENTAS  ',' PPTO','Garantías Extendidas','Alcance Plan  ',' ELEG','Conv. '],
       'SC': ['DET.','CLUB','PTTO.','Monto','Ventas','ALCANCE','ELEGIBLES','Conversión ']}
OUT = {'WM': 'Histórico WM (con estrellas).xlsx',
       'BA': 'Histórico BA (con estrellas).xlsx',
       'SC': 'Histórico SC (con estrellas).xlsx'}

def datarow(f, clave, nombre, mo, ca, me, el):
    alc = (ca / me) if me else 0
    conv = (ca / el) if el else 0
    if f == 'SC':   # DET., CLUB, PTTO.(meta), Monto($), Ventas(GE), ALCANCE, ELEGIBLES, Conversión
        return [clave, nombre, round(me), round(mo, 2), round(ca), alc, round(el), conv]
    # WM/BA: DET, TDA, VENTAS($), PPTO(meta), GE, Alcance, ELEG, Conv
    return [clave, nombre, round(mo, 2), round(me), round(ca), alc, round(el), conv]

def starval(si, i, b, pb):
    if b is None: return ''
    pend = (i == 1 and mi < DIC25)
    if pend: return ''
    if pb is None or (i == 1 and mi == DIC25): return b[i]
    now, prev = b[i], pb[i]
    if now and prev: return 1
    if not now and not prev: return 0
    return '+1' if (now and not prev) else -1

for f in ['WM', 'BA', 'SC']:
    path = os.path.join(HISTDIR, OUT[f])
    if not os.path.exists(path):
        print(f"  · aviso: no encontré {OUT[f]}, lo salto"); continue
    wb = openpyxl.load_workbook(path)
    if LABEL in wb.sheetnames: del wb[LABEL]        # idempotente
    pos = wb.sheetnames.index('Leyenda estrellas') if 'Leyenda estrellas' in wb.sheetnames else len(wb.sheetnames)
    ws = wb.create_sheet(LABEL, pos)
    hdr = HDR[f] + ['E1','E2','E3','E4','E5','Cambio nivel']
    for c, h in enumerate(hdr, 1):
        cc = ws.cell(1, c, h); cc.font = hf; cc.fill = hfill; cc.alignment = ctr
    n = 0
    for s in stores:
        if s['f'] != f: continue
        si = s['i']
        if si not in row: continue
        mo, ca, me, el = row[si]
        ws.append(datarow(f, s['clave'], s['n'], mo, ca, me, el)); n += 1
        r = ws.max_row
        ws.cell(r, 6).number_format = '0.0%'; ws.cell(r, 8).number_format = '0.0%'
        b = bits(si, k); pb = bits(si, prevk) if prevk >= 0 else None
        for i in range(5):
            ws.cell(r, 9 + i, starval(si, i, b, pb)).alignment = ctr
        L = lvl(si, k); Lp = lvl(si, prevk) if prevk >= 0 else None
        camb = '—' if (L is None or Lp is None) else ('Más' if L > Lp else ('Menos' if L < Lp else 'Igual'))
        ws.cell(r, 14, camb).alignment = ctr
    ws.column_dimensions['A'].width = 8; ws.column_dimensions['B'].width = 26
    wb.save(path)
    print(f"  · {OUT[f]}: hoja {LABEL} agregada ({n} tiendas) · hojas {len(wb.sheetnames)}")
print("Históricos (con estrellas) al día.")
